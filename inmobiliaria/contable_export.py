"""Exportación PDF y Excel de reportes contables."""

from __future__ import annotations

import io
import re
from decimal import Decimal
from typing import Any

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

from docs.services import branding_pdf_context, generar_pdf_desde_plantilla

from inmobiliaria.contable_reportes import proyecto_activo_contable


def _safe_filename(base: str) -> str:
    safe = re.sub(r"[^\w.\-]+", "_", base.strip(), flags=re.UNICODE).strip("._")
    return (safe[:100] or "reporte_contable")


def contable_pdf_context(*, pie_inmobiliaria: str, **extra: Any) -> dict[str, Any]:
    proyecto = proyecto_activo_contable()
    brand = branding_pdf_context(proyecto)
    return {
        **brand,
        "proyecto": proyecto,
        "pie_inmobiliaria": pie_inmobiliaria,
        "emitido_en": timezone.localtime(),
        **extra,
    }


def respuesta_pdf(*, template_name: str, context: dict, filename_base: str) -> HttpResponse:
    pdf_bytes = generar_pdf_desde_plantilla(template_name=template_name, context=context)
    name = _safe_filename(filename_base)
    return HttpResponse(
        pdf_bytes,
        content_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{name}.pdf"'},
    )


def _cell_val(v: Any) -> Any:
    if isinstance(v, Decimal):
        return float(v)
    return v


def _xlsx_response(wb, filename_base: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name = _safe_filename(filename_base)
    return HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}.xlsx"'},
    )


def xlsx_libro_ventas(ctx: dict) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Ventas IVA"
    ws.append(["Libro de Ventas (IVA)"])
    ws.append([f"Período: {ctx['inicio']} — {ctx['fin']}"])
    ws.append([])
    ws.append(["Fecha", "Comprobante", "Cliente", "Plan", "Concepto", "Gravado", "IVA est.", "Total"])
    for f in ctx["filas"]:
        ws.append(
            [
                f["fecha"].isoformat(),
                f["comprobante"],
                f["cliente"],
                f["contrato"],
                f["concepto"],
                _cell_val(f["gravado"]),
                _cell_val(f["iva"]),
                _cell_val(f["monto"]),
            ]
        )
    ws.append([])
    ws.append(["", "", "", "", "Totales", _cell_val(ctx["total_gravado"]), _cell_val(ctx["total_iva"]), _cell_val(ctx["total_monto"])])
    ws["A1"].font = Font(bold=True)
    return _xlsx_response(wb, f"libro_ventas_iva_{ctx['inicio'].strftime('%Y%m')}")


def xlsx_ingresos_mes(ctx: dict) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Ingresos mes"
    ws.append(["Reporte de Ingresos del Mes (Pago a Cuenta)"])
    ws.append([f"Período: {ctx['inicio']} — {ctx['fin']}"])
    ws.append([])
    ws.append(["Total recibido", _cell_val(ctx["total_recibido"])])
    ws.append(["Comprobantes", ctx["n_comprobantes"]])
    ws.append(["Cuotas (capital prog.)", _cell_val(ctx["total_cuotas"])])
    ws.append(["Abono a capital", _cell_val(ctx["total_capital"])])
    ws.append(["Intereses est.", _cell_val(ctx["total_interes"])])
    ws.append(["Recargos admin.", _cell_val(ctx["total_recargo"])])
    ws.append([])
    ws.append(["Concepto", "Monto"])
    for concepto, monto in ctx["filas_concepto"]:
        ws.append([concepto, _cell_val(monto)])
    ws["A1"].font = Font(bold=True)
    return _xlsx_response(wb, f"ingresos_mes_{ctx['inicio'].strftime('%Y%m')}")


def xlsx_cuentas_por_cobrar(ctx: dict) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Cuentas por cobrar"
    ws.append(["Cuentas por Cobrar — Antigüedad de Saldos"])
    ws.append([f"Corte: {ctx['hoy'].isoformat()}"])
    ws.append(["Total por cobrar", _cell_val(ctx["total_pendiente"])])
    ws.append([])
    ws.append(
        [
            "Antigüedad",
            "Cliente",
            "Plan",
            "Proyecto",
            "Días atraso",
            "Cuotas pend.",
            "Próx. vencimiento",
            "Saldo",
        ]
    )
    labels = {
        "al_dia": "Al día",
        "1_30": "1-30 días",
        "31_60": "31-60 días",
        "61_90": "61-90 días",
        "91_mas": "Más de 90 días",
    }
    for key, label in labels.items():
        for f in ctx["buckets"][key]:
            proy = ""
            if f["contrato"].inmueble_id and f["contrato"].inmueble.proyecto_id:
                proy = f["contrato"].inmueble.proyecto.nombre
            ws.append(
                [
                    label,
                    f"{f['cliente'].apellidos}, {f['cliente'].nombres}",
                    f["contrato"].numero,
                    proy,
                    f["dias_atraso"] or 0,
                    f["cuotas_pendientes"],
                    f["proximo_vencimiento"].isoformat(),
                    _cell_val(f["saldo_pendiente"]),
                ]
            )
    ws["A1"].font = Font(bold=True)
    return _xlsx_response(wb, f"cuentas_por_cobrar_{ctx['hoy'].strftime('%Y%m%d')}")


def xlsx_estado_capital(ctx: dict) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Capital e intereses"
    ws.append(["Estado de cuenta — Capital e Intereses"])
    ws.append([f"Período: {ctx['inicio']} — {ctx['fin']}"])
    ws.append([])
    ws.append(["Fecha", "Cliente", "Plan", "Concepto", "Capital", "Interés est.", "Recargo", "Total"])
    for f in ctx["filas"]:
        ws.append(
            [
                f["pago"].fecha.isoformat(),
                f"{f['cliente'].apellidos}, {f['cliente'].nombres}",
                f["contrato"].numero,
                f["pago"].get_concepto_display(),
                _cell_val(f["capital"]),
                _cell_val(f["interes"]),
                _cell_val(f["recargo"]),
                _cell_val(f["total"]),
            ]
        )
    ws.append([])
    ws.append(["", "", "", "Totales", _cell_val(ctx["total_capital"]), _cell_val(ctx["total_interes"]), "", _cell_val(ctx["total_recibido"])])
    ws["A1"].font = Font(bold=True)
    return _xlsx_response(wb, f"capital_intereses_{ctx['inicio'].strftime('%Y%m')}")
